import openpyxl as px


class signUpPageData:
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        sheet_name = "TestDataSheet"
        workbook = px.load_workbook("D:\\AutomatedStore\\testData\\testData.xlsx")
        sheet = workbook[sheet_name]

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value] = str(sheet.cell(row=i, column=j).value)

        return [Dict]
